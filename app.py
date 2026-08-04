import io
import json
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

from scoring import compute_all

BASE_DIR = Path(__file__).parent
STRUCTURE_PATH = BASE_DIR / "structure.json"
TEMPLATE_PATH = BASE_DIR / "template.xlsx"

st.set_page_config(page_title="Internal Audit / Visit Store", layout="wide")

st.markdown(
    """
    <style>
    /* Kategori header - meniru baris kategori berwarna di Excel */
    .cat-banner {
        background: #1f3864;
        color: #ffffff;
        font-weight: 700;
        font-size: 1.05rem;
        padding: 8px 14px;
        border-radius: 4px;
        margin: 18px 0 6px 0;
        letter-spacing: .03em;
    }
    .subcat-label {
        font-weight: 700;
        text-decoration: underline;
        margin: 10px 0 4px 4px;
        color: #1f3864;
    }
    /* Panel tombol melayang di kiri bawah */
    div.st-key-floating_actions {
        position: fixed;
        left: 18px;
        bottom: 18px;
        z-index: 9999;
        background: #ffffff;
        padding: 12px;
        border-radius: 12px;
        box-shadow: 0 4px 18px rgba(0,0,0,0.25);
        border: 1px solid #e0e0e0;
        width: 230px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# --------------------------------------------------------------------------
# Load structure & Supabase client
# --------------------------------------------------------------------------
@st.cache_data
def load_structure():
    with open(STRUCTURE_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


structure = load_structure()


@st.cache_resource
def get_supabase_client():
    try:
        from supabase import create_client
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
        return create_client(url, key)
    except Exception:
        return None


supabase = get_supabase_client()

# --------------------------------------------------------------------------
# Session state defaults
# --------------------------------------------------------------------------
defaults = {
    "store_name": "",
    "date1": "",
    "date2": "",
    "auditor": "",
    "pic_on_duty": "",
    "remarks": {},   # {"1": "text", ...}
    "loaded_key": None,
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


def reset_form():
    for k, v in defaults.items():
        st.session_state[k] = v if not isinstance(v, dict) else {}
    st.query_params.clear()


def load_draft(row: dict):
    st.session_state["store_name"] = row.get("store_name", "") or ""
    st.session_state["date1"] = row.get("audit_date", "") or ""
    st.session_state["date2"] = row.get("date2", "") or ""
    st.session_state["auditor"] = row.get("auditor", "") or ""
    st.session_state["pic_on_duty"] = row.get("pic_on_duty", "") or ""
    st.session_state["remarks"] = row.get("remarks", {}) or {}
    st.session_state["loaded_key"] = (row.get("store_name"), row.get("audit_date"))


# --------------------------------------------------------------------------
# Auto-restore from URL query params (survive browser refresh after Save)
# --------------------------------------------------------------------------
qp = st.query_params
if supabase is not None and "store" in qp and "date" in qp and st.session_state["loaded_key"] is None:
    try:
        resp = (
            supabase.table("audit_drafts")
            .select("*")
            .eq("store_name", qp["store"])
            .eq("audit_date", qp["date"])
            .limit(1)
            .execute()
        )
        if resp.data:
            load_draft(resp.data[0])
    except Exception as e:
        st.warning(f"Gagal memuat draft otomatis dari Supabase: {e}")

# --------------------------------------------------------------------------
# Header
# --------------------------------------------------------------------------
st.title("📋 Internal Audit / Visit Store")

if supabase is None:
    st.error(
        "Koneksi Supabase belum dikonfigurasi. Isi `SUPABASE_URL` dan `SUPABASE_KEY` "
        "di `.streamlit/secrets.toml` agar fitur Save/Load berfungsi. "
        "Lihat README.md."
    )

with st.expander("📂 Buka data toko yang tersimpan (lanjutkan audit)", expanded=False):
    if supabase is None:
        st.caption("Supabase belum terkoneksi.")
    else:
        try:
            resp = (
                supabase.table("audit_drafts")
                .select("store_name, audit_date, updated_at")
                .order("updated_at", desc=True)
                .limit(500)
                .execute()
            )
            rows = resp.data or []
        except Exception as e:
            rows = []
            st.warning(f"Gagal mengambil daftar draft: {e}")

        if rows:
            options = {
                f"{r['store_name']}  —  {r['audit_date']}  (update: {r['updated_at'][:16].replace('T',' ')})": r
                for r in rows
            }
            choice = st.selectbox("Pilih draft", ["-- pilih --"] + list(options.keys()))
            col_a, col_b = st.columns([1, 1])
            with col_a:
                if st.button("🔄 Muat Draft", disabled=(choice == "-- pilih --")):
                    picked = options[choice]
                    full = (
                        supabase.table("audit_drafts")
                        .select("*")
                        .eq("store_name", picked["store_name"])
                        .eq("audit_date", picked["audit_date"])
                        .limit(1)
                        .execute()
                    )
                    if full.data:
                        load_draft(full.data[0])
                        st.query_params["store"] = picked["store_name"]
                        st.query_params["date"] = picked["audit_date"]
                        st.rerun()
            with col_b:
                if st.button("🆕 Mulai Audit Baru"):
                    reset_form()
                    st.rerun()
        else:
            st.caption("Belum ada draft tersimpan.")

st.divider()

# --------------------------------------------------------------------------
# Store / header info form
# --------------------------------------------------------------------------
c1, c2 = st.columns(2)
with c1:
    st.session_state["store_name"] = st.text_input("STORE NAME", value=st.session_state["store_name"])
    st.session_state["date1"] = st.text_input("DATE", value=st.session_state["date1"], help="Sesuai kolom DATE pertama pada template Excel")
    st.session_state["date2"] = st.text_input("DATE (2)", value=st.session_state["date2"], help="Sesuai kolom DATE kedua pada template Excel")
with c2:
    st.session_state["auditor"] = st.text_input("AUDITOR", value=st.session_state["auditor"])
    st.session_state["pic_on_duty"] = st.text_input("PIC ON DUTY", value=st.session_state["pic_on_duty"])

st.divider()

# --------------------------------------------------------------------------
# Item tables (per category / subcategory) — editable Remarks column
# --------------------------------------------------------------------------
st.subheader("DETAIL TO VERIFY")
st.caption("Isi kolom **Remarks** jika ada temuan pada suatu poin. Begitu Remarks diisi, Actual Score poin tersebut otomatis menjadi 0 — persis seperti logika di file Excel.")

remarks_state = st.session_state["remarks"]


ROW_COLS = [0.5, 4.2, 0.9, 0.9, 3.5]


def _safe_remark(v) -> str:
    """Selalu kembalikan string bersih, walau v ternyata None/NaN/float."""
    if v is None:
        return ""
    if isinstance(v, float):
        return ""  # NaN or stray float -> treat as empty
    return str(v)


def render_header_row():
    h = st.columns(ROW_COLS)
    h[0].markdown("**No**")
    h[1].markdown("**Deskripsi**")
    h[2].markdown("**Basic**")
    h[3].markdown("**Actual**")
    h[4].markdown("**Remarks**")


def render_table(items: list[dict], key: str):
    render_header_row()
    for it in items:
        number = str(it["number"])
        widget_key = f"{key}_remark_{number}"

        # PENTING: kalau widget ini sudah pernah dirender, session_state sudah
        # berisi nilai TERBARU (hasil ketikan yang baru saja memicu rerun ini)
        # SEBELUM baris ini digambar ulang. Baca dari sini, bukan dari
        # remarks_state, supaya Actual Score langsung sinkron di rerun yang
        # sama -- tidak lagi telat satu langkah.
        if widget_key in st.session_state:
            current = _safe_remark(st.session_state[widget_key])
        else:
            current = _safe_remark(remarks_state.get(number, ""))

        actual = 0 if current.strip() else it["basic"]

        cols = st.columns(ROW_COLS)
        cols[0].markdown(str(it["number"]))
        cols[1].markdown(it["desc"])  # wraps naturally, no truncation
        cols[2].markdown(str(it["basic"]))
        if actual == 0:
            cols[3].markdown(f":red[**{actual}**]")
        else:
            cols[3].markdown(str(actual))

        new_val = cols[4].text_area(
            "Remarks",
            value=current,
            key=f"{key}_remark_{number}",
            label_visibility="collapsed",
            height=68,
        )
        remarks_state[number] = _safe_remark(new_val)
        st.markdown("<hr style='margin:4px 0;opacity:0.15'>", unsafe_allow_html=True)


def category_banner(name: str):
    st.markdown(f"<div class='cat-banner'>{name}</div>", unsafe_allow_html=True)


for cat in structure:
    if cat["name"] == "ETC":
        note = cat["note"]
        category_banner("ETC")
        st.info(f"{note['note_number']}. {note['note_text']}")
        continue

    category_banner(cat["name"])
    if "subcategories" in cat:
        for sub in cat["subcategories"]:
            st.markdown(f"<div class='subcat-label'>{sub['name']}</div>", unsafe_allow_html=True)
            render_table(sub["items"], key=f"editor_{sub['row']}")
    else:
        render_table(cat["items"], key=f"editor_{cat['row']}")

st.session_state["remarks"] = remarks_state

# --------------------------------------------------------------------------
# Score summary
# --------------------------------------------------------------------------
result = compute_all(structure, remarks_state)

st.divider()
st.subheader("📊 Rincian Nilai per Kategori")

summary_df = pd.DataFrame([
    {
        "Kategori": c["name"],
        "Basic Score": c["basic"],
        "Actual Score": c["actual"],
        "%": round(c["percent"], 2),
    }
    for c in result["categories"]
])
st.dataframe(summary_df, hide_index=True, width="stretch")

m1, m2, m3 = st.columns(3)
m1.metric("TOTAL SCORE (E175)", round(result["grand_percent_sum"], 2))
m2.metric("FINAL SCORE", round(result["final_score"], 2))
m3.metric("GRADING", result["grade"])

with st.expander("Keterangan Grading"):
    st.markdown(
        "- **A** = 96–100 (Excellent)\n"
        "- **B+** = 91–95 (Very Good)\n"
        "- **B** = 86–90 (Good)\n"
        "- **B-** = 81–85 (Need Improvement)\n"
        "- **C** = <80 (Critical)"
    )

st.divider()

# --------------------------------------------------------------------------
# Save for later & Download
# --------------------------------------------------------------------------
def build_excel_bytes() -> bytes:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["DETAIL AUDIT"]  # ATTACHMENT sheet is left completely untouched

    ws["B6"] = st.session_state["store_name"]
    ws["B7"] = st.session_state["date1"]
    ws["B8"] = st.session_state["date2"]
    ws["E6"] = st.session_state["auditor"]
    ws["E7"] = st.session_state["pic_on_duty"]

    for cat in structure:
        if cat["name"] == "ETC":
            continue
        item_lists = []
        if "subcategories" in cat:
            for sub in cat["subcategories"]:
                item_lists.append(sub["items"])
        else:
            item_lists.append(cat["items"])
        for items in item_lists:
            for it in items:
                r = remarks_state.get(str(it["number"]), "") or ""
                actual = 0 if r.strip() else it["basic"]
                ws.cell(it["row"], 4).value = actual  # column D
                ws.cell(it["row"], 5).value = r if r.strip() else None  # column E

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


st.markdown("<div style='height:140px'></div>", unsafe_allow_html=True)  # spacer agar tidak ketutup panel melayang

with st.container(key="floating_actions"):
    if st.button("💾 Save for Later", width="stretch"):
        if supabase is None:
            st.error("Supabase belum terkoneksi, tidak bisa menyimpan.")
        elif not st.session_state["store_name"] or not st.session_state["date1"]:
            st.error("STORE NAME dan DATE wajib diisi sebelum menyimpan.")
        else:
            payload = {
                "store_name": st.session_state["store_name"],
                "audit_date": st.session_state["date1"],
                "date2": st.session_state["date2"],
                "auditor": st.session_state["auditor"],
                "pic_on_duty": st.session_state["pic_on_duty"],
                "remarks": remarks_state,
                "updated_at": datetime.utcnow().isoformat(),
            }
            try:
                supabase.table("audit_drafts").upsert(
                    payload, on_conflict="store_name,audit_date"
                ).execute()
                st.query_params["store"] = st.session_state["store_name"]
                st.query_params["date"] = st.session_state["date1"]
                st.session_state["loaded_key"] = (st.session_state["store_name"], st.session_state["date1"])
                st.success("Tersimpan.")
            except Exception as e:
                st.error(f"Gagal menyimpan: {e}")

    excel_bytes = build_excel_bytes()
    fname = f"Audit_{st.session_state['store_name'] or 'Store'}_{st.session_state['date1'] or 'Date'}.xlsx".replace(" ", "_")
    st.download_button(
        "⬇️ Download Excel",
        data=excel_bytes,
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )
