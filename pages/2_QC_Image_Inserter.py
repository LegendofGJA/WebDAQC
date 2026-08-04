import streamlit as st
import os
import shutil
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from PIL import Image as PILImage
from io import BytesIO
from style import (
    inject_css, inject_sidebar_brand, inject_footer,
    log_traffic, correct_orientation, extract_datetime,
    get_preset_files, convert_date_to_english,
)
from audit_core import list_saved_drafts, fetch_draft, build_filled_workbook


# ── Helpers ──────────────────────────────────────────────

def col_letter_to_num(letter):
    result = 0
    for ch in letter.strip().upper():
        if ch.isalpha():
            result = result * 26 + (ord(ch) - 64)
    return result


def col_num_to_letter(num):
    result = ""
    while num > 0:
        num, remainder = divmod(num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def parse_cell_ref(cell_str):
    cell_str = cell_str.strip().upper()
    col_part = ""
    row_part = ""
    for ch in cell_str:
        if ch.isalpha():
            col_part += ch
        elif ch.isdigit():
            row_part += ch
    col_num = col_letter_to_num(col_part) if col_part else None
    row_num = int(row_part) if row_part else None
    return col_num, row_num


# ── Page Config ──────────────────────────────────────────

st.set_page_config(page_title="QC Image Inserter", page_icon="📸", layout="wide")
inject_css()
inject_sidebar_brand()

# PAGE HEADER
st.markdown(
    """
    <div class="page-head">
        <div class="page-head-icon">📸</div>
        <h2>QC Image Inserter</h2>
        <p>Susun foto QC lapangan ke dalam template Excel secara otomatis</p>
    </div>
    """,
    unsafe_allow_html=True,
)

# USER INPUT
st.markdown(
    '<div class="ilp-user"><div class="ilp-user-icon">👤</div>'
    '<div class="ilp-user-hint">Masukkan nama Anda untuk memulai</div></div>',
    unsafe_allow_html=True,
)
user_name = st.text_input("Nama Pengguna", placeholder="Ketik nama Anda di sini...")
if not user_name:
    st.warning("Silakan isi Nama Pengguna di atas agar sistem bisa digunakan.")
    st.stop()

# STEP INDICATOR
st.markdown(
    """
    <div class="ilp-steps">
        <div class="ilp-step active"><div class="ilp-step-num">1</div><div class="ilp-step-label">Info Lokasi</div></div>
        <div class="ilp-step-line active"></div>
        <div class="ilp-step active"><div class="ilp-step-num">2</div><div class="ilp-step-label">Template</div></div>
        <div class="ilp-step-line active"></div>
        <div class="ilp-step active"><div class="ilp-step-num">3</div><div class="ilp-step-label">Upload Foto</div></div>
        <div class="ilp-step-line"></div>
        <div class="ilp-step"><div class="ilp-step-num">4</div><div class="ilp-step-label">Export</div></div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ── ROW 1: INFO LOKASI & TEMPLATE EXCEL ─────────────────

col_info, col_template = st.columns(2, gap="medium")

# --- Template dulu dihitung, supaya kalau pilih draft, info toko bisa
# auto-terisi dari draft sebelum kartu Info Lokasi dirender di kolom kiri ---
draft_data = None
use_draft = False

with col_template:
    st.markdown(
        """
        <div class="card-head">
            <div class="card-head-icon">📊</div>
            <h3>Template Excel</h3>
            <p>Upload, pilih preset, atau ambil dari draft DETAIL AUDIT</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    mode_template = st.radio(
        "Sumber template:",
        ["Dari Draft DETAIL AUDIT", "File Preset", "Upload Manual"],
        horizontal=True,
    )

    excel_file = None
    preset_name = ""
    selected_sheet = None
    all_sheet_names = []

    if mode_template == "Dari Draft DETAIL AUDIT":
        use_draft = True
        drafts = list_saved_drafts()
        if drafts:
            options = {
                f"{r['store_name']} — {r['audit_date']}": r for r in drafts
            }
            pick = st.selectbox("Pilih draft tersimpan:", ["-- pilih --"] + list(options.keys()))
            if pick != "-- pilih --":
                picked = options[pick]
                draft_data = fetch_draft(picked["store_name"], picked["audit_date"])
                if draft_data:
                    st.success(f"Draft dimuat: {picked['store_name']} — {picked['audit_date']}")
                    preset_name = f"Draft: {picked['store_name']} {picked['audit_date']}"
                    all_sheet_names = ["DETAIL AUDIT", "ATTACHMENT"]
                    selected_sheet = "ATTACHMENT"  # foto selalu ke sheet ini
        else:
            st.warning("Belum ada draft DETAIL AUDIT tersimpan. Isi & simpan dulu di halaman DETAIL AUDIT.")

    elif mode_template == "File Preset":
        available_presets = get_preset_files()
        if available_presets:
            preset_pilihan = st.selectbox("Pilih file preset:", available_presets)
            preset_name = preset_pilihan
            preset_path = os.path.join("presets", preset_pilihan)
            with open(preset_path, "rb") as f:
                excel_file = BytesIO(f.read())
        else:
            st.warning("Folder 'presets/' kosong.")
    else:
        excel_file = st.file_uploader("Pilih file Excel (.xlsx)", type=["xlsx"], label_visibility="collapsed")
        preset_name = "Upload Manual"

    if excel_file:
        try:
            excel_file.seek(0)
            wb_scan = load_workbook(excel_file, read_only=True)
            all_sheet_names = wb_scan.sheetnames
            wb_scan.close()
            default_sheet = "ATTACHMENT" if "ATTACHMENT" in all_sheet_names else all_sheet_names[0]
            selected_sheet = st.selectbox(
                "Target Sheet (Foto):",
                all_sheet_names,
                index=all_sheet_names.index(default_sheet) if default_sheet in all_sheet_names else 0,
            )
        except Exception as e:
            st.error(f"Gagal membaca struktur Excel: {e}")

with col_info:
    st.markdown(
        """
        <div class="card-head">
            <div class="card-head-icon">📍</div>
            <h3>Informasi Lokasi QC</h3>
            <p>Detail lokasi untuk penamaan file output</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if use_draft and draft_data:
        nama_toko = draft_data.get("store_name", "")
        tanggal_qc = draft_data.get("audit_date", "")
        pic = draft_data.get("pic_on_duty", "")
        st.info(
            f"**Toko:** {nama_toko}  \n**Tanggal:** {tanggal_qc}  \n"
            f"**Auditor (draft):** {draft_data.get('auditor','')}  \n**PIC:** {pic}"
        )
        st.caption("Data ini otomatis diambil dari draft DETAIL AUDIT, tidak perlu diketik ulang.")
    else:
        nama_toko = st.text_input("Nama Toko / Area", placeholder="Contoh: Batavia PIK")
        tanggal_qc = st.text_input("Tanggal QC", placeholder="Contoh: 10 Mar 26")
        pic = st.text_input("PIC", placeholder="Contoh: Budi Santoso")

# ── CELL WRITING MODE ────────────────────────────────────

st.markdown("---")
if use_draft:
    st.caption("Cell B6/B7/E6/E7 di sheet DETAIL AUDIT sudah otomatis terisi dari draft — tidak perlu diatur manual.")
    cell_mode = "Auto"
    cell_target_sheet = "DETAIL AUDIT"
else:
    st.markdown("##### Store Name, Audit & Date QC")
    cell_mode = st.radio(
        "Pilih mode pengisian cell Excel (B6, B7, E6, E7):",
        ["Manual", "Auto"],
        horizontal=True,
        help="Manual: Anda isi sendiri di Excel setelah download. Auto: Sistem otomatis isi dari data yang sudah dimasukkan.",
    )

    cell_target_sheet = None
    if cell_mode == "Auto":
        if all_sheet_names:
            default_cell_sheet = "DETAIL AUDIT" if "DETAIL AUDIT" in all_sheet_names else all_sheet_names[0]
            cell_target_sheet = st.selectbox(
                "Target Sheet (Cell B6/B7/E6/E7):",
                all_sheet_names,
                index=all_sheet_names.index(default_cell_sheet) if default_cell_sheet in all_sheet_names else 0,
                help="Pilih sheet tempat menulis Store Name, Tanggal, Nama Pengguna, dan PIC.",
            )
            st.markdown(
                f"""<div style="background:rgba(229,50,45,0.06); border:1px solid rgba(229,50,45,0.15);
                     border-radius:10px; padding:12px 16px; margin-top:8px;">
                    <p style="color:#fca5a5; font-size:0.82rem; margin:0; line-height:1.6;">
                        Cell <b>B6</b> = Nama Toko &nbsp;|&nbsp;
                        Cell <b>B7</b> = Tanggal QC &nbsp;|&nbsp;
                        Cell <b>E6</b> = Nama Pengguna &nbsp;|&nbsp;
                        Cell <b>E7</b> = PIC<br>
                        Target: <b>{cell_target_sheet}</b>
                    </p>
                </div>""",
                unsafe_allow_html=True,
            )
        else:
            st.warning("Tidak ada sheet yang terdeteksi.")

# ── PENGATURAN LAYOUT QC PHOTOS ──────────────────────────

st.markdown("---")
st.markdown(
    """
    <div class="card-head">
        <div class="card-head-icon">⚙️</div>
        <h3>Pengaturan Layout</h3>
        <p>Atur ukuran dan posisi gambar di Excel</p>
    </div>
    """,
    unsafe_allow_html=True,
)

layout_option = st.selectbox("Opsi Layout:", ["Default", "Custom"])

if layout_option == "Default":
    ROWS = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30]
    COLS = [1, 2, 3, 4, 5, 6]
    COL_W = 20.43
    ROW_H = 123.75
    IMAGE_WIDTH_CM = 3.2
    IMAGE_HEIGHT_CM = 4.10
else:
    c_row, c_col = st.columns(2)
    r_in = c_row.text_input("Rows (koma)", "2,4,6,8,10,12")
    c_in = c_col.text_input("Columns (huruf, koma)", "A,B,C,D,E,F")
    try:
        ROWS = [int(x.strip()) for x in r_in.split(",")]
        COLS = [col_letter_to_num(x.strip()) for x in c_in.split(",")]
    except Exception:
        st.error("Format salah! Rows pakai angka (2,4,6), Columns pakai huruf (A,B,C)")
        st.stop()
    c_w, c_h = st.columns(2)
    COL_W = c_w.number_input("Col Width", value=20.43)
    ROW_H = c_h.number_input("Row Height", value=123.75)
    i_w, i_h = st.columns(2)
    IMAGE_WIDTH_CM = i_w.number_input("Image Width (cm)", value=3.2)
    IMAGE_HEIGHT_CM = i_h.number_input("Image Height (cm)", value=4.10)

# ── TANDA TANGAN — PENGESEHAN FORM QC ────────────────────

st.markdown("---")
st.markdown(
    """
    <div class="card-head">
        <div class="card-head-icon">✍️</div>
        <h3>Pengesahan Form QC</h3>
        <p>Upload foto tanda tangan Auditor dan PIC sebagai pengesahan form quality control</p>
    </div>
    """,
    unsafe_allow_html=True,
)

ttd_mode = st.selectbox("Opsi Tanda Tangan:", ["Default", "Custom"])

if ttd_mode == "Default":
    TTD_IMG_AUDITOR_CELL = "A161"
    TTD_IMG_PIC_CELL = "D161"
    TTD_NAMA_AUDITOR_CELL = "A163"
    TTD_NAMA_PIC_CELL = "D163"
    TTD_IMG_W = 3.20
    TTD_IMG_H = 3.20
    TTD_COL_W = 19.71
    TTD_ROW_H = 97.5
else:
    st.markdown("##### Posisi Cell Tanda Tangan")
    ca, cb = st.columns(2)
    TTD_IMG_AUDITOR_CELL = ca.text_input("Cell gambar TTD Auditor", value="A161")
    TTD_IMG_PIC_CELL = cb.text_input("Cell gambar TTD PIC", value="D161")
    cc, cd = st.columns(2)
    TTD_NAMA_AUDITOR_CELL = cc.text_input("Cell nama Auditor", value="A163")
    TTD_NAMA_PIC_CELL = cd.text_input("Cell nama PIC", value="D163")

    st.markdown("##### Ukuran Gambar & Cell TTD")
    ce, cf = st.columns(2)
    TTD_IMG_W = ce.number_input("Image Width (cm)", value=3.20, step=0.1, format="%.2f")
    TTD_IMG_H = cf.number_input("Image Height (cm)", value=3.20, step=0.1, format="%.2f")
    cg, c_row_h = st.columns(2)
    TTD_COL_W = cg.number_input("Col Width", value=19.71, step=0.1, format="%.2f")
    TTD_ROW_H = c_row_h.number_input("Row Height", value=97.5, step=0.5, format="%.1f")

col_ttd_auditor, col_ttd_pic = st.columns(2, gap="medium")

with col_ttd_auditor:
    st.markdown(
        """
        <div class="card-head">
            <h3>Tanda Tangan Auditor</h3>
            <p>Nama dari Nama Pengguna</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    ttd_auditor = st.file_uploader(
        "Upload foto tanda tangan Auditor",
        type=["jpg", "jpeg", "png", "webp"],
        key="ttd_auditor",
    )
    if ttd_auditor:
        st.success(f"TTD Auditor: {ttd_auditor.name}")

with col_ttd_pic:
    st.markdown(
        """
        <div class="card-head">
            <h3>Tanda Tangan PIC</h3>
            <p>Nama dari input PIC</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    ttd_pic = st.file_uploader(
        "Upload foto tanda tangan PIC",
        type=["jpg", "jpeg", "png", "webp"],
        key="ttd_pic",
    )
    if ttd_pic:
        st.success(f"TTD PIC: {ttd_pic.name}")

# ── UPLOAD FOTO QC ───────────────────────────────────────

st.markdown("---")
st.markdown(
    """
    <div class="upload-head">
        <div class="upload-head-icon">📸</div>
        <h3>Upload Foto QC</h3>
        <p>Pilih semua foto sekaligus dari galeri HP</p>
    </div>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """<div style="background:rgba(245,158,11,0.08); border:1px solid rgba(245,158,11,0.2);
         border-radius:10px; padding:10px 14px; margin-bottom:14px;">
        <p style="color:#fcd34d; font-size:0.78rem; margin:0; line-height:1.5;">
            <b>Tips HP:</b> Gunakan Gallery bawaan HP atau File Manager.
            Google Photos sering memutus koneksi. Jika reconnecting,
            tunggu beberapa detik lalu coba lagi.
        </p>
    </div>""",
    unsafe_allow_html=True,
)

if "uploader_key" not in st.session_state:
    st.session_state.uploader_key = 0

if st.button("Hapus Semua Foto (Reset)"):
    st.session_state.uploader_key += 1
    st.rerun()

uploaded_photos = st.file_uploader(
    "Pilih semua foto sekaligus dari Galeri",
    type=["jpg", "jpeg", "png", "webp"],
    accept_multiple_files=True,
    key=f"photos_{st.session_state.uploader_key}",
    label_visibility="collapsed",
)

if uploaded_photos:
    st.success(f"{len(uploaded_photos)} foto dipilih")
else:
    st.caption("Belum ada foto yang dipilih")

# ── ACTION ───────────────────────────────────────────────

st.markdown("---")
st.markdown(
    """
    <div class="ilp-action-head">
        <h3>Siap untuk memproses?</h3>
        <p>Pastikan semua data sudah benar sebelum mengeksekusi</p>
    </div>
    """,
    unsafe_allow_html=True,
)

if st.button("MULAI EXPORT DAN PROSES DATA", type="primary", use_container_width=True):
    if not nama_toko or not tanggal_qc:
        st.warning("Silakan isi Nama Toko dan Tanggal QC!")
    elif use_draft and not draft_data:
        st.warning("Silakan pilih draft DETAIL AUDIT yang mau dipakai!")
    elif not use_draft and not excel_file:
        st.warning("Silakan upload atau pilih Excel Template Master!")
    elif not uploaded_photos:
        st.warning("Silakan pilih foto QC!")
    elif not selected_sheet:
        st.error("Target sheet foto tidak valid.")
    elif cell_mode == "Auto" and not cell_target_sheet:
        st.error("Target sheet cell tidak valid.")
    else:
        with st.spinner("Sedang memproses dan mengompres foto..."):
            try:
                template_str = f"{preset_name}, {selected_sheet}"
                log_traffic(user_name, nama_toko, tanggal_qc, template_str, layout_option)

                if use_draft and draft_data:
                    wb = build_filled_workbook(
                        draft_data.get("store_name", ""),
                        draft_data.get("audit_date", ""),
                        draft_data.get("date2", ""),
                        draft_data.get("auditor", ""),
                        draft_data.get("pic_on_duty", ""),
                        draft_data.get("remarks", {}) or {},
                    )
                else:
                    excel_file.seek(0)
                    wb = load_workbook(excel_file)

                temp_dir = "temp_web_photos"
                os.makedirs(temp_dir, exist_ok=True)

                # ── Tentukan sheet untuk penulisan cell & tanda tangan ──
                if cell_mode == "Auto" and cell_target_sheet:
                    write_sheet_name = cell_target_sheet
                else:
                    write_sheet_name = (
                        "DETAIL AUDIT" if "DETAIL AUDIT" in wb.sheetnames
                        else wb.sheetnames[0]
                    )

                # ── AUTO: Tulis B6, B7, E6, E7 (skip kalau dari draft, sudah terisi) ──
                if not use_draft and cell_mode == "Auto" and write_sheet_name in wb.sheetnames:
                    ws_cell = wb[write_sheet_name]
                    ws_cell["B6"] = nama_toko
                    ws_cell["B7"] = convert_date_to_english(tanggal_qc)
                    ws_cell["E6"] = user_name
                    ws_cell["E7"] = pic

                # ── TANDA TANGAN ──
                if write_sheet_name in wb.sheetnames:
                    ws_ttd = wb[write_sheet_name]

                    # Set kolom width & row height untuk cell gambar TTD
                    for cell_ref in [TTD_IMG_AUDITOR_CELL, TTD_IMG_PIC_CELL]:
                        col_num, row_num = parse_cell_ref(cell_ref)
                        if col_num and row_num:
                            col_letter = col_num_to_letter(col_num)
                            ws_ttd.column_dimensions[col_letter].width = TTD_COL_W
                            ws_ttd.row_dimensions[row_num].height = TTD_ROW_H

                    # Tulis nama Auditor & PIC dengan alignment center
                    cell_auditor = ws_ttd[TTD_NAMA_AUDITOR_CELL]
                    cell_auditor.value = user_name
                    cell_auditor.alignment = Alignment(horizontal='center', vertical='center')

                    cell_pic = ws_ttd[TTD_NAMA_PIC_CELL]
                    cell_pic.value = pic
                    cell_pic.alignment = Alignment(horizontal='center', vertical='center')

                    # TTD Auditor → gambar
                    if ttd_auditor:
                        with PILImage.open(ttd_auditor) as img_pil:
                            img_pil = correct_orientation(img_pil)
                            if img_pil.mode in ("RGBA", "P"):
                                img_pil = img_pil.convert("RGB")
                            img_pil.thumbnail((800, 400), PILImage.Resampling.LANCZOS)
                            ttd_a_path = os.path.join(temp_dir, "ttd_auditor.jpg")
                            img_pil.save(ttd_a_path, format="JPEG", quality=85, optimize=True)
                        img_ttd_a = ExcelImage(ttd_a_path)
                        img_ttd_a.width = int(TTD_IMG_W * 37.8)
                        img_ttd_a.height = int(TTD_IMG_H * 37.8)
                        ws_ttd.add_image(img_ttd_a, TTD_IMG_AUDITOR_CELL)

                    # TTD PIC → gambar
                    if ttd_pic:
                        with PILImage.open(ttd_pic) as img_pil:
                            img_pil = correct_orientation(img_pil)
                            if img_pil.mode in ("RGBA", "P"):
                                img_pil = img_pil.convert("RGB")
                            img_pil.thumbnail((800, 400), PILImage.Resampling.LANCZOS)
                            ttd_p_path = os.path.join(temp_dir, "ttd_pic.jpg")
                            img_pil.save(ttd_p_path, format="JPEG", quality=85, optimize=True)
                        img_ttd_p = ExcelImage(ttd_p_path)
                        img_ttd_p.width = int(TTD_IMG_W * 37.8)
                        img_ttd_p.height = int(TTD_IMG_H * 37.8)
                        ws_ttd.add_image(img_ttd_p, TTD_IMG_PIC_CELL)

                # ── FOTO QC: Susun ke sheet target ──
                ws = wb[selected_sheet]

                for c in COLS:
                    ws.column_dimensions[col_num_to_letter(c)].width = COL_W
                for r in ROWS:
                    ws.row_dimensions[r].height = ROW_H

                sorted_photos = sorted(
                    uploaded_photos,
                    key=lambda x: extract_datetime(x.name, x),
                    reverse=True,
                )
                all_cells = [f"{col_num_to_letter(col)}{row}" for row in ROWS for col in COLS]
                success_count = 0

                for i in range(min(len(sorted_photos), len(all_cells))):
                    photo = sorted_photos[i]
                    temp_path = os.path.join(temp_dir, f"compressed_img_{i}.jpg")

                    with PILImage.open(photo) as img_pil:
                        img_pil = correct_orientation(img_pil)
                        if img_pil.mode in ("RGBA", "P"):
                            img_pil = img_pil.convert("RGB")
                        img_pil.thumbnail((1280, 1280), PILImage.Resampling.LANCZOS)
                        img_pil.save(
                            temp_path, format="JPEG", quality=82,
                            optimize=True, subsampling=0,
                        )

                    img_excel = ExcelImage(temp_path)
                    img_excel.width = int(IMAGE_WIDTH_CM * 37.8)
                    img_excel.height = int(IMAGE_HEIGHT_CM * 37.8)
                    ws.add_image(img_excel, all_cells[i])
                    success_count += 1

                output = BytesIO()
                wb.save(output)
                wb.close()
                output.seek(0)
                shutil.rmtree(temp_dir, ignore_errors=True)

                final_filename = f"{nama_toko.strip()} {tanggal_qc.strip()}.xlsx"

                cell_msg = ""
                if cell_mode == "Auto" and cell_target_sheet:
                    cell_msg = (
                        f" | Cell B6/B7/E6/E7 terisi otomatis"
                        f" di sheet '{cell_target_sheet}'"
                    )
                st.success(f"Berhasil menyusun {success_count} foto!{cell_msg}")

                st.download_button(
                    label="DOWNLOAD FILE EXCEL",
                    data=output,
                    file_name=final_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"Terjadi kesalahan: {e}")

inject_footer()
