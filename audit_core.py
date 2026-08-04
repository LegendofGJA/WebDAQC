"""
Modul inti yang dipakai bersama oleh halaman DETAIL AUDIT dan QC Image
Inserter: koneksi Supabase, simpan/ambil/hapus draft, dan generator
workbook DETAIL AUDIT yang sudah terisi (dipakai ulang sebagai basis
sebelum foto QC ditempel di halaman kedua).
"""
from __future__ import annotations

import io
import json
from pathlib import Path

import openpyxl
import streamlit as st

BASE_DIR = Path(__file__).parent
STRUCTURE_PATH = BASE_DIR / "structure.json"
TEMPLATE_PATH = BASE_DIR / "template.xlsx"

DRAFT_TABLE = "audit_drafts"


@st.cache_data
def load_structure():
    with open(STRUCTURE_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


@st.cache_resource
def get_supabase_client():
    try:
        from supabase import create_client
        url = st.secrets["SUPABASE_URL"]
        key = st.secrets["SUPABASE_KEY"]
        return create_client(url, key)
    except Exception:
        return None


def list_saved_drafts(limit: int = 500):
    """Daftar ringkas semua draft tersimpan, terbaru dulu."""
    supabase = get_supabase_client()
    if supabase is None:
        return []
    try:
        resp = (
            supabase.table(DRAFT_TABLE)
            .select("store_name, audit_date, updated_at")
            .order("updated_at", desc=True)
            .limit(limit)
            .execute()
        )
        return resp.data or []
    except Exception:
        return []


def fetch_draft(store_name: str, audit_date: str):
    supabase = get_supabase_client()
    if supabase is None:
        return None
    try:
        resp = (
            supabase.table(DRAFT_TABLE)
            .select("*")
            .eq("store_name", store_name)
            .eq("audit_date", audit_date)
            .limit(1)
            .execute()
        )
        return resp.data[0] if resp.data else None
    except Exception:
        return None


def save_draft(payload: dict):
    supabase = get_supabase_client()
    if supabase is None:
        raise RuntimeError("Supabase belum terkoneksi.")
    return (
        supabase.table(DRAFT_TABLE)
        .upsert(payload, on_conflict="store_name,audit_date")
        .execute()
    )


def delete_draft(store_name: str, audit_date: str):
    supabase = get_supabase_client()
    if supabase is None:
        raise RuntimeError("Supabase belum terkoneksi.")
    return (
        supabase.table(DRAFT_TABLE)
        .delete()
        .eq("store_name", store_name)
        .eq("audit_date", audit_date)
        .execute()
    )


def build_filled_workbook(
    store_name: str,
    date1: str,
    date2: str,
    auditor: str,
    pic_on_duty: str,
    remarks: dict,
) -> openpyxl.Workbook:
    """
    Load template.xlsx bersih, isi header (B6/B7/B8/E6/E7) + Actual
    Score/Remarks tiap item sesuai `remarks`. Mengembalikan objek
    Workbook (BUKAN bytes) supaya bisa terus diproses (mis. ditambah
    foto QC di sheet ATTACHMENT) sebelum akhirnya di-save.
    """
    structure = load_structure()
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["DETAIL AUDIT"]

    ws["B6"] = store_name
    ws["B7"] = date1
    ws["B8"] = date2
    ws["E6"] = auditor
    ws["E7"] = pic_on_duty

    for cat in structure:
        if cat["name"] == "ETC":
            continue
        item_lists = (
            [s["items"] for s in cat["subcategories"]]
            if "subcategories" in cat
            else [cat["items"]]
        )
        for items in item_lists:
            for it in items:
                r = (remarks.get(str(it["number"]), "") or "").strip()
                actual = 0 if r else it["basic"]
                ws.cell(it["row"], 4).value = actual
                ws.cell(it["row"], 5).value = r if r else None

    return wb


def workbook_to_bytes(wb: openpyxl.Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
